# -*- coding: utf-8 -*-
"""
Stock Engine Pro — v4.1
Multi-method stock analysis: Valuation · Fundamentals · Technicals · Sentiment

v4.1 changes:
  - Removed username/password system (single-user, no login required)
  - Fixed database update issues (WAL mode, per-operation connections,
    schema migration via ALTER TABLE ADD COLUMN)
  - Added Database & Methodology tab with Sarbanes-Oxley certification
  - Fixed pandas Styler.applymap → Styler.map (deprecated in pandas ≥ 2.1)
  - Fixed safe column access throughout UI
  - Tuned valuation UNDERVALUED threshold (4 → 3)
  - Guard _detail accesses against cold-load KeyErrors
"""
import io
import os
import re
import time
import datetime
import sqlite3
import urllib.request
import xml.etree.ElementTree as ET
from email.utils import parsedate_to_datetime as _parse_rfc2822

import numpy as np
import pandas as pd
import yfinance as yf
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =============================================================================
# CONFIGURATION
# =============================================================================

DB_FILENAME   = "stock_pro_v7.db"
TEMPLATE_PATH = "Updated_Workbook_Template.xlsx"
STALE_HOURS   = 24

SECTOR_BENCHMARKS = {
    "Technology":             {"PE": 30, "PS": 6.0, "PB": 8.0, "EV_EBITDA": 20},
    "Healthcare":             {"PE": 25, "PS": 4.0, "PB": 4.0, "EV_EBITDA": 15},
    "Financial Services":     {"PE": 14, "PS": 3.0, "PB": 1.5, "EV_EBITDA": 10},
    "Energy":                 {"PE": 10, "PS": 1.5, "PB": 1.8, "EV_EBITDA":  6},
    "Consumer Cyclical":      {"PE": 20, "PS": 2.5, "PB": 4.0, "EV_EBITDA": 14},
    "Industrials":            {"PE": 20, "PS": 2.0, "PB": 3.5, "EV_EBITDA": 12},
    "Utilities":              {"PE": 18, "PS": 2.5, "PB": 2.0, "EV_EBITDA": 10},
    "Consumer Defensive":     {"PE": 22, "PS": 2.0, "PB": 4.0, "EV_EBITDA": 15},
    "Real Estate":            {"PE": 35, "PS": 6.0, "PB": 3.0, "EV_EBITDA": 18},
    "Communication Services": {"PE": 20, "PS": 4.0, "PB": 3.0, "EV_EBITDA": 12},
    "Basic Materials":        {"PE": 15, "PS": 1.5, "PB": 2.0, "EV_EBITDA":  8},
}
DEFAULT_BENCHMARKS = {"PE": 20, "PS": 3.0, "PB": 3.0, "EV_EBITDA": 12}

POSITIVE_WORDS = {
    "beat","beats","surge","surges","soar","soars","record","rally",
    "gain","gains","strong","growth","profit","upgrade","bullish",
    "outperform","exceed","exceeds","buy","positive","upside","boost",
}
NEGATIVE_WORDS = {
    "miss","misses","plunge","plunges","fall","falls","drop","drops",
    "loss","losses","weak","decline","downgrade","bearish","underperform",
    "disappoint","disappoints","sell","negative","downside","cut","risk",
}

# Full list of columns in the analysis table (used for schema migration)
ANALYSIS_COLUMNS = {
    "Ticker":              "TEXT PRIMARY KEY",
    "Price":               "REAL",
    "Verdict_Overall":     "TEXT",
    "Verdict_Technical":   "TEXT",
    "Verdict_Fundamental": "TEXT",
    "Verdict_Valuation":   "TEXT",
    "Score_Tech":          "INTEGER",
    "Score_Fund":          "INTEGER",
    "Score_Val":           "INTEGER",
    "Sector":              "TEXT",
    "PE_Ratio":            "REAL",
    "Forward_PE":          "REAL",
    "PEG_Ratio":           "REAL",
    "PS_Ratio":            "REAL",
    "PB_Ratio":            "REAL",
    "EV_EBITDA":           "REAL",
    "Graham_Number":       "REAL",
    "Intrinsic_Value":     "REAL",
    "DCF_Value":           "REAL",
    "Beta":                "REAL",
    "Earnings_Beat_Rate":  "REAL",
    "Volume_Signal":       "TEXT",
    "Profit_Margins":      "REAL",
    "ROE":                 "REAL",
    "Debt_to_Equity":      "REAL",
    "RSI":                 "REAL",
    "MACD_Signal":         "TEXT",
    "SMA_Status":          "TEXT",
    "Sentiment_Score":     "REAL",
    "Sentiment_Label":     "TEXT",
    "Last_Updated":        "TEXT",
}


# =============================================================================
# 1. UTILITIES
# =============================================================================

def safe_num(value):
    """Convert value to float, returning None on failure."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return None if (np.isnan(value) or np.isinf(value)) else float(value)
        except Exception:
            return None
    if isinstance(value, str):
        if value.lower() in {"n/a", "none", "nan", "inf"}:
            return None
        try:
            return float(value.replace("%", "").replace(",", "").strip())
        except ValueError:
            return None
    return None


def fmt(value, fmt_str=".2f", prefix="", suffix="", scale=1.0, fallback="N/A"):
    v = safe_num(value)
    if v is None:
        return fallback
    return f"{prefix}{v * scale:{fmt_str}}{suffix}"


def row_get(row, key, default=None):
    """Safe accessor for both pandas Series and plain dicts."""
    try:
        val = row[key]
        return default if (val is None or (isinstance(val, float) and np.isnan(val))) else val
    except (KeyError, TypeError):
        return default


def get_color(verdict):
    v = str(verdict).upper()
    if "STRONG BUY"  in v: return "#00C853"
    if "BUY"         in v: return "#69F0AE"
    if "STRONG SELL" in v: return "#D50000"
    if "SELL"        in v: return "#FF5252"
    return "#9E9E9E"


def _fin_val(df, keys, col):
    if df is None or df.empty:
        return None
    for key in keys:
        if key in df.index:
            try:
                return safe_num(df.loc[key, col])
            except Exception:
                pass
    return None


def _retry(fn, attempts=3, delay=2.0):
    for i in range(attempts):
        try:
            result = fn()
            if result is not None:
                return result
        except Exception:
            pass
        if i < attempts - 1:
            time.sleep(delay)
    return None


def _fetch_rss_headlines(ticker):
    url = f"https://finance.yahoo.com/rss/headline?s={ticker}"
    try:
        req = urllib.request.Request(
            url, headers={"User-Agent": "Mozilla/5.0 (compatible; StockEngine/4.1)"}
        )
        with urllib.request.urlopen(req, timeout=6) as resp:
            xml_data = resp.read()
        root     = ET.fromstring(xml_data)
        now      = datetime.datetime.now(datetime.timezone.utc)
        items    = []
        for item in root.findall(".//item"):
            title    = (item.findtext("title") or "").strip()
            pub_raw  = item.findtext("pubDate") or ""
            age_days = 7.0
            if pub_raw:
                try:
                    pub_dt   = _parse_rfc2822(pub_raw)
                    age_days = max(0.0, (now - pub_dt).total_seconds() / 86400)
                except Exception:
                    pass
            if title:
                items.append({"title": title, "age_days": age_days})
        return items
    except Exception:
        return []


def _hours_since(timestamp_str):
    try:
        dt = datetime.datetime.strptime(str(timestamp_str), "%Y-%m-%d %H:%M")
        return (datetime.datetime.now() - dt).total_seconds() / 3600
    except Exception:
        return 0.0


def _styler_map(styler, fn, subset=None):
    """pandas ≥2.1 renamed Styler.applymap → Styler.map. Handle both."""
    try:
        return styler.map(fn, subset=subset)
    except AttributeError:
        return styler.applymap(fn, subset=subset)


# =============================================================================
# 2. DATABASE MANAGER  (single-user, no auth)
# =============================================================================

class DatabaseManager:
    """
    SQLite manager with:
      - WAL journal mode for reliable concurrent reads/writes
      - Per-operation connections to avoid stale state
      - Automatic schema migration (ALTER TABLE ADD COLUMN) for new fields
    """

    def __init__(self, db_path: str):
        self.db_path = os.path.abspath(db_path)
        self._init_schema()

    # -- Connection helpers ---------------------------------------------------

    def _conn(self):
        """Open a fresh autocommit connection with WAL mode."""
        conn = sqlite3.connect(self.db_path, isolation_level=None,
                               check_same_thread=False)
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.row_factory = sqlite3.Row
        return conn

    # -- Schema creation + migration ------------------------------------------

    def _init_schema(self):
        with self._conn() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS watchlist (
                    ticker     TEXT PRIMARY KEY,
                    added_date TEXT
                );
                CREATE TABLE IF NOT EXISTS analysis (
                    Ticker TEXT PRIMARY KEY,
                    Last_Updated TEXT
                );
                CREATE TABLE IF NOT EXISTS analysis_history (
                    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                    Ticker              TEXT,
                    run_at              TEXT,
                    Price               REAL,
                    Verdict_Overall     TEXT,
                    Verdict_Valuation   TEXT,
                    Verdict_Fundamental TEXT,
                    Verdict_Technical   TEXT,
                    Sentiment_Label     TEXT,
                    Score_Val           INTEGER,
                    Score_Fund          INTEGER,
                    Score_Tech          INTEGER,
                    Sentiment_Score     REAL,
                    DCF_Value           REAL,
                    Graham_Number       REAL
                );
            """)
        # Migrate analysis table: add any missing columns
        self._migrate_analysis_columns()

    def _migrate_analysis_columns(self):
        """Add any columns from ANALYSIS_COLUMNS that don't exist yet."""
        with self._conn() as conn:
            existing = {
                row[1] for row in conn.execute("PRAGMA table_info(analysis);")
            }
            for col, col_type in ANALYSIS_COLUMNS.items():
                if col not in existing and "PRIMARY KEY" not in col_type:
                    try:
                        conn.execute(
                            f"ALTER TABLE analysis ADD COLUMN {col} "
                            f"{col_type.replace(' PRIMARY KEY', '')};"
                        )
                    except sqlite3.OperationalError:
                        pass  # Column already exists (race condition)

    # -- Watchlist (single-user) ----------------------------------------------

    def toggle_watchlist(self, ticker: str) -> str:
        with self._conn() as conn:
            exists = conn.execute(
                "SELECT 1 FROM watchlist WHERE ticker=?", (ticker,)
            ).fetchone()
            if exists:
                conn.execute("DELETE FROM watchlist WHERE ticker=?", (ticker,))
                return "removed"
            else:
                date = datetime.datetime.now().strftime("%Y-%m-%d")
                conn.execute(
                    "INSERT INTO watchlist (ticker, added_date) VALUES (?, ?)",
                    (ticker, date),
                )
                return "added"

    def is_on_watchlist(self, ticker: str) -> bool:
        with self._conn() as conn:
            return conn.execute(
                "SELECT 1 FROM watchlist WHERE ticker=?", (ticker,)
            ).fetchone() is not None

    def get_watchlist(self) -> pd.DataFrame:
        with self._conn() as conn:
            return pd.read_sql_query(
                """SELECT w.ticker, a.Price, a.Verdict_Overall
                   FROM watchlist w
                   LEFT JOIN analysis a ON w.ticker = a.Ticker
                   ORDER BY w.added_date DESC""",
                conn,
            )

    def get_watchlist_comparison(self) -> pd.DataFrame:
        with self._conn() as conn:
            return pd.read_sql_query(
                """SELECT w.ticker,
                          a.Price, a.Sector,
                          a.Verdict_Overall,
                          a.Verdict_Valuation,  a.Score_Val,
                          a.Verdict_Fundamental,a.Score_Fund,
                          a.Verdict_Technical,  a.Score_Tech,
                          a.Sentiment_Label,    a.Sentiment_Score,
                          a.PE_Ratio, a.PB_Ratio, a.EV_EBITDA,
                          a.RSI, a.Beta,
                          a.DCF_Value, a.Graham_Number,
                          a.Earnings_Beat_Rate,
                          a.Last_Updated
                   FROM watchlist w
                   LEFT JOIN analysis a ON w.ticker = a.Ticker
                   ORDER BY w.added_date DESC""",
                conn,
            )

    # -- Analysis cache -------------------------------------------------------

    def save_analysis(self, data: dict):
        """Upsert latest snapshot and append to history."""
        keys   = list(data.keys())
        values = list(data.values())
        sql    = (
            f"INSERT OR REPLACE INTO analysis ({', '.join(keys)}) "
            f"VALUES ({', '.join(['?']*len(keys))})"
        )
        with self._conn() as conn:
            conn.execute(sql, values)

            # History row
            HIST_COLS = [
                "Ticker","Price","Verdict_Overall","Verdict_Valuation",
                "Verdict_Fundamental","Verdict_Technical","Sentiment_Label",
                "Score_Val","Score_Fund","Score_Tech","Sentiment_Score",
                "DCF_Value","Graham_Number",
            ]
            hist_vals = [data.get(c) for c in HIST_COLS]
            run_at    = data.get("Last_Updated", "")
            conn.execute(
                f"INSERT INTO analysis_history (run_at, {', '.join(HIST_COLS)}) "
                f"VALUES ({', '.join(['?']*(len(HIST_COLS)+1))})",
                [run_at] + hist_vals,
            )

    def get_analysis(self, ticker: str) -> pd.DataFrame:
        with self._conn() as conn:
            return pd.read_sql_query(
                "SELECT * FROM analysis WHERE Ticker=?",
                conn, params=(ticker,),
            )

    def get_analysis_history(self, ticker: str) -> pd.DataFrame:
        with self._conn() as conn:
            return pd.read_sql_query(
                "SELECT * FROM analysis_history WHERE Ticker=? ORDER BY run_at ASC",
                conn, params=(ticker,),
            )

    def get_all_tickers(self) -> list:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT DISTINCT Ticker FROM analysis ORDER BY Ticker"
            ).fetchall()
            return [r[0] for r in rows]


# =============================================================================
# 3. ANALYST ENGINE
# =============================================================================

class StockAnalyst:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def _fetch(self, ticker):
        def _try():
            stock = yf.Ticker(ticker)
            hist  = stock.history(period="1y")
            info  = stock.info
            if hist.empty or not info or not (info.get("longName") or info.get("shortName")):
                return None
            news = stock.news or []
            return hist, info, news, stock
        result = _retry(_try, attempts=3, delay=2.0)
        if result is None:
            return None, None, [], None
        return result

    def validate_ticker(self, ticker):
        def _try():
            info = yf.Ticker(ticker.strip().upper()).info
            name = info.get("longName") or info.get("shortName")
            return name if name else None
        name = _retry(_try, attempts=2, delay=1.5)
        if name:
            return True, name
        return False, (
            f"'{ticker}' was not recognised. "
            "Check the symbol (e.g. AAPL, MSFT, NVDA) and try again."
        )

    # -- DCF ------------------------------------------------------------------
    def _calc_dcf(self, stock_obj, info, price):
        try:
            cf = stock_obj.cashflow
            if cf is None or cf.empty:
                return None, "No cash-flow data available"
            dates    = sorted(cf.columns.tolist())[-3:]
            fcf_vals = []
            for dt in dates:
                op_cf = _fin_val(cf, ["Operating Cash Flow","Cash From Operations",
                                      "Total Cash From Operating Activities"], dt)
                capex = _fin_val(cf, ["Capital Expenditure","Capital Expenditures",
                                      "Purchase Of Property Plant And Equipment",
                                      "Purchases Of Property Plant And Equipment"], dt)
                if op_cf is not None:
                    fcf_vals.append(op_cf - abs(capex or 0))
            if not fcf_vals:
                return None, "Insufficient FCF data"
            base_fcf = sum(fcf_vals) / len(fcf_vals)
            if base_fcf <= 0:
                return None, f"Negative avg FCF (${base_fcf/1e9:.1f}B) — DCF not applicable"
            raw_g = safe_num(info.get("earningsGrowth")) or safe_num(info.get("revenueGrowth")) or 0.08
            g1    = max(-0.05, min(raw_g, 0.25))
            g2    = g1 * 0.5
            g_tv  = 0.025
            beta  = max(0.5, min(safe_num(info.get("beta")) or 1.0, 3.0))
            wacc  = max(0.07, min(0.04 + beta * 0.055, 0.18))
            pv, fcf = 0.0, base_fcf
            for yr in range(1, 4):
                fcf *= (1 + g1);  pv += fcf / (1 + wacc) ** yr
            for yr in range(4, 6):
                fcf *= (1 + g2);  pv += fcf / (1 + wacc) ** yr
            tv    = fcf * (1 + g_tv) / (wacc - g_tv)
            pv_tv = tv / (1 + wacc) ** 5
            ev    = pv + pv_tv
            debt  = safe_num(info.get("totalDebt")) or 0
            cash  = safe_num(info.get("totalCash")) or 0
            shares = safe_num(info.get("sharesOutstanding")) or 1
            equity = ev - debt + cash
            if equity <= 0:
                return None, "Negative equity value after debt bridge"
            iv = equity / shares
            return iv, (
                f"DCF: g₁={g1*100:.1f}% · g₂={g2*100:.1f}% · "
                f"WACC={wacc*100:.1f}% · TV={pv_tv/ev*100:.0f}% of EV"
            )
        except Exception as exc:
            return None, f"DCF error: {exc}"

    # -- Technical ------------------------------------------------------------
    def _score_technical(self, hist, info):
        close  = hist["Close"]
        volume = hist["Volume"]
        delta  = close.diff()
        gain   = delta.where(delta > 0, 0).rolling(14).mean()
        loss   = (-delta.where(delta < 0, 0)).rolling(14).mean()
        rsi    = (100 - 100 / (1 + gain / loss.replace(0, np.nan))).iloc[-1]
        sma50  = close.rolling(50).mean().iloc[-1]
        sma200 = close.rolling(200).mean().iloc[-1]
        price  = close.iloc[-1]
        macd   = close.ewm(span=12, adjust=False).mean() - close.ewm(span=26, adjust=False).mean()
        sig    = macd.ewm(span=9,  adjust=False).mean()
        macd_bull  = bool(macd.iloc[-1] > sig.iloc[-1])
        avg_vol20  = volume.rolling(20).mean().iloc[-1]
        recent_vol = volume.iloc[-5:].mean()
        vol_ratio  = recent_vol / avg_vol20 if avg_vol20 > 0 else 1.0
        vol_spike  = bool(vol_ratio > 1.25)
        vol_signal = "Normal"
        beta       = safe_num(info.get("beta"))

        score, reasons = 0, []
        if price > sma200:
            score += 1; reasons.append("Price > 200-SMA ✅")
        else:
            score -= 1; reasons.append("Price < 200-SMA ❌")

        if sma50 > sma200:
            score += 1; reasons.append("Golden Cross (50 > 200 SMA) ✅")
        else:
            reasons.append("Death Cross (50 < 200 SMA) ❌")

        if rsi < 30:
            score += 2; reasons.append(f"RSI {rsi:.1f} — Oversold ✅")
        elif rsi > 70:
            score -= 2; reasons.append(f"RSI {rsi:.1f} — Overbought ❌")
        else:
            reasons.append(f"RSI {rsi:.1f} — Neutral")

        if macd_bull:
            score += 1; reasons.append("MACD above Signal Line ✅")
        else:
            score -= 1; reasons.append("MACD below Signal Line ❌")

        if vol_spike:
            if price > sma50:
                score += 1; vol_signal = "High-Volume Breakout"
                reasons.append(f"High-volume breakout ({vol_ratio:.1f}× 20d avg) ✅")
            else:
                score -= 1; vol_signal = "High-Volume Breakdown"
                reasons.append(f"High-volume breakdown ({vol_ratio:.1f}× 20d avg) ❌")
        else:
            reasons.append(f"Volume normal ({vol_ratio:.1f}× 20-day avg)")

        if beta is not None:
            vol_desc = (
                "High volatility" if beta > 1.5
                else "Low volatility" if beta < 0.7
                else "Market-like volatility"
            )
            reasons.append(f"Beta {beta:.2f} — {vol_desc}")

        return {
            "verdict":     "BUY" if score >= 3 else ("SELL" if score <= -1 else "HOLD"),
            "score":       score, "rsi": rsi,
            "sma_status":  "Bullish" if price > sma200 else "Bearish",
            "macd_signal": "Bullish" if macd_bull else "Bearish",
            "price": price, "beta": beta,
            "vol_signal": vol_signal, "vol_ratio": vol_ratio, "reasons": reasons,
        }

    # -- Earnings surprises ---------------------------------------------------
    def _score_earnings_surprises(self, stock_obj):
        try:
            eh = None
            for attr in ("earnings_history", "get_earnings_history"):
                obj = getattr(stock_obj, attr, None)
                eh  = obj() if callable(obj) else obj
                if eh is not None and not (hasattr(eh, "empty") and eh.empty):
                    break
            if eh is None or (hasattr(eh, "empty") and eh.empty):
                return 0, "No earnings history available", None
            recent = eh.tail(4)
            beats = misses = 0
            for _, q in recent.iterrows():
                actual = safe_num(q.get("epsActual") or q.get("EPS Actual") or q.get("Reported EPS"))
                est    = safe_num(q.get("epsEstimate") or q.get("EPS Estimate") or q.get("EPS Estimated"))
                if actual is not None and est is not None:
                    if actual > est:   beats  += 1
                    elif actual < est: misses += 1
            total = beats + misses
            if total == 0:
                return 0, "Earnings comparison unavailable", None
            rate = beats / total
            if rate >= 0.75:
                return 2, f"Beat EPS estimates {beats}/{total} recent quarters ✅✅", rate
            elif rate >= 0.50:
                return 1, f"Beat EPS estimates {beats}/{total} recent quarters ✅", rate
            elif rate <= 0.25:
                return -1, f"Missed EPS estimates {misses}/{total} recent quarters ❌", rate
            else:
                return 0, f"Mixed earnings results ({beats}/{total} beats)", rate
        except Exception:
            return 0, "Earnings data unavailable", None

    # -- Fundamental ----------------------------------------------------------
    def _score_fundamental(self, info, stock_obj):
        roe     = safe_num(info.get("returnOnEquity"))
        margins = safe_num(info.get("profitMargins"))
        debt_eq = safe_num(info.get("debtToEquity"))
        score, reasons = 0, []
        if roe is not None:
            if roe > 0.15: score += 1; reasons.append(f"ROE {roe*100:.1f}% > 15% ✅")
            else:          reasons.append(f"ROE {roe*100:.1f}% < 15% ❌")
        if margins is not None:
            if margins > 0.20: score += 1; reasons.append(f"Net Margin {margins*100:.1f}% > 20% ✅")
            else:              reasons.append(f"Net Margin {margins*100:.1f}% < 20% ❌")
        if debt_eq is not None:
            if   debt_eq < 100: score += 1; reasons.append(f"D/E {debt_eq:.0f}% < 100% ✅")
            elif debt_eq > 200: score -= 1; reasons.append(f"D/E {debt_eq:.0f}% > 200% — High Risk ❌")
            else:               reasons.append(f"D/E {debt_eq:.0f}% — Moderate")
        es_score, es_msg, beat_rate = self._score_earnings_surprises(stock_obj)
        score += es_score
        if es_msg:
            reasons.append(es_msg)
        verdict = "STRONG" if score >= 3 else ("STABLE" if score >= 1 else "WEAK")
        return {
            "verdict": verdict, "score": score,
            "roe": roe, "margins": margins, "debt_eq": debt_eq,
            "beat_rate": beat_rate, "reasons": reasons,
        }

    # -- Valuation ------------------------------------------------------------
    def _score_valuation(self, info, price, dcf_value):
        sector = info.get("sector", "Unknown")
        bench  = SECTOR_BENCHMARKS.get(sector, DEFAULT_BENCHMARKS)
        pe         = safe_num(info.get("trailingPE"))
        forward_pe = safe_num(info.get("forwardPE"))
        peg        = safe_num(info.get("pegRatio"))
        ps         = safe_num(info.get("priceToSalesTrailing12Months"))
        pb         = safe_num(info.get("priceToBook"))
        ev_ebitda  = safe_num(info.get("enterpriseToEbitda"))
        eps        = safe_num(info.get("trailingEps"))
        bvps       = safe_num(info.get("bookValue"))
        score, reasons = 0, []

        if pe is not None:
            if pe < bench["PE"]: score += 1; reasons.append(f"P/E {pe:.1f} < sector {bench['PE']} ✅")
            else:                reasons.append(f"P/E {pe:.1f} ≥ sector {bench['PE']} ❌")
        if ev_ebitda is not None:
            if ev_ebitda < bench["EV_EBITDA"]:
                score += 1; reasons.append(f"EV/EBITDA {ev_ebitda:.1f} < {bench['EV_EBITDA']} ✅")
            else:
                reasons.append(f"EV/EBITDA {ev_ebitda:.1f} ≥ {bench['EV_EBITDA']} ❌")
        if pb is not None:
            if pb < bench["PB"]: score += 1; reasons.append(f"P/B {pb:.1f} < sector {bench['PB']} ✅")
            else:                reasons.append(f"P/B {pb:.1f} ≥ sector {bench['PB']} ❌")
        if peg is not None and peg > 0:
            if peg < 1.0: score += 1; reasons.append(f"PEG {peg:.2f} < 1.0 — Undervalued growth ✅")
            else:         reasons.append(f"PEG {peg:.2f} ≥ 1.0")

        graham = None
        if eps and bvps and eps > 0 and bvps > 0:
            graham = (22.5 * eps * bvps) ** 0.5
            if price < graham:
                score += 1; reasons.append(f"Price ${price:.2f} < Graham ${graham:.2f} ✅")
            elif price > graham * 1.5:
                score -= 1; reasons.append(f"Price ${price:.2f} > 1.5× Graham ${graham:.2f} ❌")
            else:
                reasons.append(f"Price ${price:.2f} near Graham ${graham:.2f}")

        if dcf_value is not None and dcf_value > 0:
            mos = (dcf_value - price) / dcf_value
            if mos >= 0.25:
                score += 2; reasons.append(f"Price ${price:.2f} — {mos*100:.0f}% below DCF ${dcf_value:.2f} (strong MoS) ✅✅")
            elif mos >= 0:
                score += 1; reasons.append(f"Price ${price:.2f} below DCF ${dcf_value:.2f} ✅")
            elif mos >= -0.15:
                reasons.append(f"Price ${price:.2f} slightly above DCF ${dcf_value:.2f}")
            else:
                score -= 1; reasons.append(f"Price ${price:.2f} — {-mos*100:.0f}% above DCF ${dcf_value:.2f} ❌")

        # Threshold: >= 3 for UNDERVALUED (tuned from 4 to be more realistic)
        verdict = "UNDERVALUED" if score >= 3 else ("FAIR VALUE" if score >= 1 else "OVERVALUED")
        return {
            "verdict": verdict, "score": score, "sector": sector, "bench": bench,
            "pe": pe, "forward_pe": forward_pe, "peg": peg, "ps": ps,
            "pb": pb, "ev_ebitda": ev_ebitda, "graham": graham, "reasons": reasons,
        }

    # -- Sentiment ------------------------------------------------------------
    def _score_sentiment(self, news, ticker):
        rss_items = _fetch_rss_headlines(ticker)
        source    = "Yahoo Finance RSS"
        if rss_items:
            raw_pool = [{"title": it["title"], "age_days": it["age_days"]} for it in rss_items[:30]]
        else:
            source   = "yfinance news feed"
            raw_pool = []
            for item in news[:15]:
                title = ""
                if isinstance(item, dict):
                    if "content" in item and isinstance(item["content"], dict):
                        title = item["content"].get("title", "")
                    else:
                        title = item.get("title", "")
                if not title:
                    continue
                age_days = 3.0
                try:
                    pub_ts = (
                        item.get("providerPublishTime")
                        or (item.get("content") or {}).get("pubDate", 0)
                    )
                    if pub_ts and isinstance(pub_ts, (int, float)):
                        age_days = max(0.0, (time.time() - pub_ts) / 86400)
                except Exception:
                    pass
                raw_pool.append({"title": title, "age_days": age_days})

        articles = []
        for it in raw_pool:
            title    = it["title"]
            age_days = it["age_days"]
            decay    = float(np.exp(-0.08 * age_days))
            words    = set(re.sub(r"[^a-z ]", "", title.lower()).split())
            raw_s    = len(words & POSITIVE_WORDS) - len(words & NEGATIVE_WORDS)
            articles.append({
                "title": title, "raw": raw_s, "decay": decay,
                "weighted": raw_s * decay,
                "signal":   "🟢 Positive" if raw_s > 0 else ("🔴 Negative" if raw_s < 0 else "⚪ Neutral"),
                "age_days": int(age_days),
            })

        if not articles:
            return {"score": 0.0, "label": "NEUTRAL", "articles": [], "reasons": [], "source": source}

        pos = sum(1 for a in articles if a["raw"] > 0)
        neg = sum(1 for a in articles if a["raw"] < 0)
        net = float(np.clip(sum(a["weighted"] for a in articles) / len(articles), -1.0, 1.0))
        label = (
            "BULLISH"        if net >= 0.15  else
            "MILDLY BULLISH" if net >  0     else
            "BEARISH"        if net <= -0.15 else
            "MILDLY BEARISH" if net <  0     else
            "NEUTRAL"
        )
        return {
            "score": net, "label": label, "articles": articles, "source": source,
            "reasons": [
                f"Scored {len(articles)} headlines ({pos} positive, {neg} negative)",
                f"Recency-weighted score: {net:+.3f}  (exponential decay, half-life ≈8.7 days)",
                f"Source: {source}",
            ],
        }

    # -- Composite verdict ----------------------------------------------------
    def _composite_verdict(self, val_v, fund_v, tech_v, sent_label):
        val_sig  = {"UNDERVALUED": 1, "FAIR VALUE": 0, "OVERVALUED": -1}[val_v]
        fund_sig = {"STRONG": 1,  "STABLE": 0, "WEAK": -1}[fund_v]
        tech_sig = {"BUY":   1,   "HOLD":  0, "SELL": -1}[tech_v]
        sent_sig = 1 if "BULLISH" in sent_label else (-1 if "BEARISH" in sent_label else 0)
        c = 0.35 * val_sig + 0.25 * fund_sig + 0.25 * tech_sig + 0.15 * sent_sig
        if   c >= 0.55:  return "STRONG BUY"
        elif c >= 0.20:  return "BUY"
        elif c <= -0.55: return "STRONG SELL"
        elif c <= -0.20: return "SELL"
        else:            return "HOLD"

    # -- Main entry point -----------------------------------------------------
    def analyze(self, ticker):
        ticker = ticker.strip().upper()
        hist, info, news, stock_obj = self._fetch(ticker)
        if hist is None:
            return None
        price                   = hist["Close"].iloc[-1]
        dcf_value, dcf_note     = self._calc_dcf(stock_obj, info, price)
        tech  = self._score_technical(hist, info)
        fund  = self._score_fundamental(info, stock_obj)
        val   = self._score_valuation(info, price, dcf_value)
        sent  = self._score_sentiment(news, ticker)
        final = self._composite_verdict(
            val["verdict"], fund["verdict"], tech["verdict"], sent["label"]
        )
        record = {
            "Ticker":              ticker,        "Price":               price,
            "Verdict_Overall":     final,
            "Verdict_Technical":   tech["verdict"],
            "Verdict_Fundamental": fund["verdict"],
            "Verdict_Valuation":   val["verdict"],
            "Score_Tech":          tech["score"],  "Score_Fund": fund["score"],
            "Score_Val":           val["score"],   "Sector":     val["sector"],
            "PE_Ratio":            val["pe"],       "Forward_PE": val["forward_pe"],
            "PEG_Ratio":           val["peg"],      "PS_Ratio":   val["ps"],
            "PB_Ratio":            val["pb"],       "EV_EBITDA":  val["ev_ebitda"],
            "Graham_Number":       val["graham"] or 0.0,
            "Intrinsic_Value":     dcf_value or val["graham"] or 0.0,
            "DCF_Value":           dcf_value or 0.0,
            "Beta":                tech["beta"],
            "Earnings_Beat_Rate":  fund["beat_rate"],
            "Volume_Signal":       tech["vol_signal"],
            "Profit_Margins":      fund["margins"],
            "ROE":                 fund["roe"],
            "Debt_to_Equity":      fund["debt_eq"],
            "RSI":                 tech["rsi"],
            "MACD_Signal":         tech["macd_signal"],
            "SMA_Status":          tech["sma_status"],
            "Sentiment_Score":     sent["score"],
            "Sentiment_Label":     sent["label"],
            "Last_Updated":        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        st.session_state["_detail"] = {
            "tech": tech, "fund": fund, "val": val, "sent": sent,
            "hist": hist, "info": info, "dcf_note": dcf_note, "dcf_value": dcf_value,
        }
        st.session_state["_xlsx_bytes"]  = None
        st.session_state["_xlsx_ticker"] = None
        self.db.save_analysis(record)
        return record


# =============================================================================
# 4. WORKBOOK GENERATOR
# =============================================================================

def _cs(cell, bold=False, size=11, color="000000", bg=None,
        align="left", wrap=False, num_fmt=None, italic=False):
    cell.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt


def generate_workbook(ticker, record, info, hist, detail):
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found at '{TEMPLATE_PATH}'")

    wb           = load_workbook(TEMPLATE_PATH)
    company_name = info.get("longName", ticker)
    sector       = record.get("Sector", "Unknown")
    overall      = record.get("Verdict_Overall", "HOLD")

    vs = wb["Valuation"]
    vs["B1"] = company_name;  vs["B2"] = ticker
    vs["C16"] = record["Price"]
    vs["C20"] = "Global"
    vs["C21"] = datetime.datetime.now().strftime("%Y-%m-%d")
    vs["C22"] = "Stock Engine Pro v4.1"
    vs["C23"] = "Initiation"
    vs["C24"] = "Buy" if "BUY" in overall else ("Sell" if "SELL" in overall else "Hold")

    try:
        so = yf.Ticker(ticker)
        af, ab, ac = so.financials, so.balance_sheet, so.cashflow

        def asc(df):
            return sorted(df.columns.tolist()) if df is not None and not df.empty else []

        fd, bd, cd = asc(af), asc(ab), asc(ac)
        n = min(5, len(fd))
        if fd:
            wb["IS"]["C6"] = fd[0].year

        hw = wb["Historical"];  dp = wb["DuPont"]
        COLS = ["C","D","E","F","G","H"]
        for i in range(n):
            cl = COLS[i]
            if i < len(fd):
                dt  = fd[i]
                rev  = _fin_val(af, ["Total Revenue","Revenue"], dt)
                ebit = _fin_val(af, ["EBIT","Operating Income"], dt)
                ni   = _fin_val(af, ["Net Income","Net Income Common Stockholders"], dt)
                gp   = _fin_val(af, ["Gross Profit"], dt)
                pre  = _fin_val(af, ["Pretax Income"], dt)
                tax  = _fin_val(af, ["Tax Provision"], dt)
                if rev:             hw[f"{cl}7"]  = rev
                if ebit:            hw[f"{cl}8"]  = ebit
                if ni:              hw[f"{cl}12"] = ni
                if gp   and rev:    dp[f"{cl}22"] = gp / rev
                if ebit and rev:    dp[f"{cl}23"] = ebit / rev
                if pre  and tax and pre != 0:
                    dp[f"{cl}26"] = tax / pre
            if i < len(bd):
                dt = bd[i]
                for keys, cell in [
                    (["Long Term Debt","Long Term Debt And Capital Lease Obligation"], f"{cl}13"),
                    (["Cash And Cash Equivalents","Cash Cash Equivalents And Short Term Investments"], f"{cl}14"),
                    (["Stockholders Equity","Total Equity Gross Minority Interest","Common Stock Equity"], f"{cl}15"),
                    (["Ordinary Shares Number"], f"{cl}17"),
                ]:
                    v = _fin_val(ab, keys, dt)
                    if v: hw[cell] = v
            if i < len(cd):
                dt  = cd[i]
                dna = _fin_val(ac, ["Depreciation And Amortization","Reconciled Depreciation"], dt)
                if dna: hw[f"{cl}9"] = abs(dna)

        ph = so.history(period="5y")
        if not ph.empty:
            ph.index   = pd.to_datetime(ph.index)
            ph["Year"] = ph.index.year
            apx        = ph.groupby("Year").agg({"High":"max","Low":"min"})
            for i, cl in enumerate(["C","D","E","F","G"]):
                if i < len(apx):
                    yr = sorted(apx.index)[i]
                    hw[f"{cl}28"] = apx.loc[yr,"High"]
                    hw[f"{cl}36"] = apx.loc[yr,"Low"]
    except Exception:
        pass

    # Summary sheet
    ss   = wb.create_sheet("Stock Engine Summary", 0)
    DARK = "1F2937"; MID = "374151"; LGRAY = "F3F4F6"; WHITE = "FFFFFF"
    G_BG = "D1FAE5"; G_FG = "065F46"
    R_BG = "FEE2E2"; R_FG = "991B1B"
    A_BG = "FEF3C7"; A_FG = "92400E"

    def vc(v):
        u = str(v).upper()
        if any(x in u for x in ("STRONG BUY","UNDERVALUED","STRONG","BULLISH")): return G_BG, G_FG
        if any(x in u for x in ("STRONG SELL","OVERVALUED","WEAK","BEARISH")):   return R_BG, R_FG
        if any(x in u for x in ("BUY","MILDLY BULLISH")):  return "C6F6D5","276749"
        if any(x in u for x in ("SELL","MILDLY BEARISH")): return "FED7D7","9B2C2C"
        return A_BG, A_FG

    for col, w in [("A",3),("B",32),("C",18),("D",18),("E",18),("F",18),("G",32)]:
        ss.column_dimensions[col].width = w

    r = 2
    ss.merge_cells(f"B{r}:G{r}")
    ss[f"B{r}"].value = f"{company_name}   ·   {ticker}   ·   {sector}"
    _cs(ss[f"B{r}"], bold=True, size=15, color=WHITE, bg=DARK, align="center")
    ss.row_dimensions[r].height = 30; r += 1

    for bc, lbl, val, nf in [
        ("B","Analysis Date",record.get("Last_Updated",""),None),
        ("D","Current Price",record["Price"],"$#,##0.00"),
        ("F","Sector",sector,None),
    ]:
        nc = chr(ord(bc)+1)
        ss[f"{bc}{r}"].value = lbl; _cs(ss[f"{bc}{r}"], bold=True, color=WHITE, bg=MID)
        ss[f"{nc}{r}"].value = val; _cs(ss[f"{nc}{r}"], color=WHITE, bg=MID, num_fmt=nf)
    ss.row_dimensions[r].height = 18; r += 1

    v_bg, v_fg = vc(overall)
    ss.merge_cells(f"B{r}:G{r}")
    ss[f"B{r}"].value = f"VERDICT:  {overall}"
    _cs(ss[f"B{r}"], bold=True, size=18, color=v_fg, bg=v_bg, align="center")
    ss.row_dimensions[r].height = 40; r += 2

    bench = SECTOR_BENCHMARKS.get(sector, DEFAULT_BENCHMARKS)
    for lbl, col in [("Engine","B"),("Verdict","C"),("Score","D"),("Weight","E"),("Notes","F")]:
        ss[f"{col}{r}"].value = lbl
        _cs(ss[f"{col}{r}"], bold=True, color=WHITE, bg=MID, align="center")
    ss.row_dimensions[r].height = 18; r += 1

    dcf_v  = safe_num(record.get("DCF_Value"))
    for eng, verdict, score, weight, note in [
        ("Valuation",    record.get("Verdict_Valuation","N/A"),
         record.get("Score_Val","–"), "35%", f"P/E<{bench['PE']}  Graham  DCF"),
        ("Fundamentals", record.get("Verdict_Fundamental","N/A"),
         record.get("Score_Fund","–"), "25%", "ROE  Margin  D/E  EPS Surprises"),
        ("Technicals",   record.get("Verdict_Technical","N/A"),
         record.get("Score_Tech","–"), "25%", "RSI  MACD  SMA  Volume  Beta"),
        ("Sentiment",    record.get("Sentiment_Label","N/A"),
         f"{safe_num(record.get('Sentiment_Score',0)) or 0:+.2f}", "15%",
         "RSS headlines + recency decay"),
    ]:
        e_bg, e_fg = vc(verdict)
        for val, col in [(eng,"B"),(verdict,"C"),(score,"D"),(weight,"E"),(note,"F")]:
            ss[f"{col}{r}"].value = val
            _cs(ss[f"{col}{r}"], bold=(col=="C"), color=e_fg, bg=e_bg,
                align="center" if col in ("C","D","E") else "left")
        ss.row_dimensions[r].height = 18; r += 1

    graham  = safe_num(record.get("Graham_Number"))
    price_v = record["Price"]
    if dcf_v and dcf_v > 0:
        r += 1
        ss.merge_cells(f"B{r}:G{r}")
        g_str = f"${graham:,.2f}" if graham else "N/A"
        ss[f"B{r}"].value = (
            f"Intrinsic Value —  DCF: ${dcf_v:,.2f}  |  Graham: {g_str}  |  "
            f"Market: ${price_v:,.2f}  |  DCF MoS: {(dcf_v-price_v)/dcf_v*100:.1f}%"
        )
        bg, fg = vc("BUY") if dcf_v > price_v else vc("SELL")
        _cs(ss[f"B{r}"], bold=True, size=11, color=fg, bg=bg, align="center")
        ss.row_dimensions[r].height = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# =============================================================================
# 5. STREAMLIT APP
# =============================================================================

st.set_page_config(page_title="Stock Engine Pro", layout="wide", page_icon="📈")

# -- Bootstrap session state --------------------------------------------------
if "db"  not in st.session_state:
    st.session_state.db      = DatabaseManager(DB_FILENAME)
    st.session_state.analyst = StockAnalyst(st.session_state.db)

db  = st.session_state.db
bot = st.session_state.analyst

# -- SIDEBAR: Watchlist -------------------------------------------------------
with st.sidebar:
    st.title("📈 Stock Engine Pro")
    st.divider()
    st.subheader("⭐ Watchlist")
    wl = db.get_watchlist()
    if not wl.empty:
        for _, wrow in wl.iterrows():
            c1, c2 = st.columns([5, 1])
            v = row_get(wrow, "Verdict_Overall") or "–"
            c1.markdown(
                f"**{wrow['ticker']}** — "
                f"<span style='color:{get_color(v)}'>{v}</span>",
                unsafe_allow_html=True,
            )
            if c2.button("✖", key=f"rm_{wrow['ticker']}"):
                db.toggle_watchlist(wrow["ticker"]); st.rerun()
    else:
        st.caption("No tickers in your watchlist yet.")

# -- MAIN ---------------------------------------------------------------------
st.title("📈 Stock Engine Pro")
st.markdown("### Multi-Method Analysis: Technicals · Fundamentals · Valuation · Sentiment")

# Input row
col_input, col_btn = st.columns([3, 1])
with col_input:
    ticker_input = st.text_input(
        "Ticker Symbol (e.g. AAPL, NVDA, MSFT)", ""
    ).strip().upper()
with col_btn:
    st.write(""); st.write("")
    run_clicked = st.button("🔍 Run Full Analysis", type="primary", use_container_width=True)

# Validation + analysis run
if run_clicked and ticker_input:
    with st.spinner(f"Validating {ticker_input}…"):
        is_valid, name_or_err = bot.validate_ticker(ticker_input)
    if not is_valid:
        st.error(f"❌ {name_or_err}")
    else:
        with st.spinner(f"Running all engines on **{name_or_err}** ({ticker_input})…"):
            result = bot.analyze(ticker_input)
        if not result:
            st.error(
                f"Analysis failed for **{ticker_input}** after 3 retries. "
                "The data provider may be temporarily unavailable — please try again."
            )

if not ticker_input:
    st.info("Enter a ticker symbol above and click **Run Full Analysis** to get started.")
    st.stop()

# Load data from DB
df = db.get_analysis(ticker_input)
if df.empty:
    st.info("No analysis found. Press **Run Full Analysis** above.")
    st.stop()

row    = df.iloc[0]
detail = st.session_state.get("_detail", {})   # may be empty on cold load

# Staleness warning
hours_old = _hours_since(str(row_get(row, "Last_Updated", "")))
if hours_old > STALE_HOURS:
    st.warning(
        f"⚠️ Cached data is **{hours_old:.0f} hours old** "
        f"(last updated {row_get(row, 'Last_Updated')}). "
        "Re-run the analysis for current prices and signals."
    )

# Top-level tabs
page_analysis, page_portfolio, page_history, page_methodology = st.tabs(
    ["🔍 Analysis", "📋 Portfolio", "📈 Score History", "📚 Database & Methodology"]
)


# =============================================================================
# TAB 1 — ANALYSIS
# =============================================================================
with page_analysis:
    st.divider()
    h1, h2, h3 = st.columns([1, 2, 1])
    with h1:
        price_val = row_get(row, "Price", 0)
        st.metric("Current Price", f"${price_val:,.2f}")
        st.caption(f"Last updated: {row_get(row, 'Last_Updated', 'N/A')}")
    with h2:
        overall_v = row_get(row, "Verdict_Overall", "N/A")
        color = get_color(overall_v)
        st.markdown(
            f"<h2 style='text-align:center; color:{color};'>⚡ {overall_v}</h2>",
            unsafe_allow_html=True,
        )
        st.markdown(
            f"<p style='text-align:center; color:#888;'>Sector: {row_get(row, 'Sector', 'N/A')}</p>",
            unsafe_allow_html=True,
        )
    with h3:
        on_wl  = db.is_on_watchlist(ticker_input)
        wl_lbl = "⭐ Remove from Watchlist" if on_wl else "☆ Add to Watchlist"
        if st.button(wl_lbl, use_container_width=True):
            db.toggle_watchlist(ticker_input); st.rerun()

    # Score badges
    st.divider()
    b1, b2, b3, b4 = st.columns(4)
    for col, (label, verdict, score) in zip(
        [b1, b2, b3, b4],
        [
            ("💰 Valuation",    row_get(row, "Verdict_Valuation",    "N/A"), row_get(row, "Score_Val")),
            ("🏢 Fundamentals", row_get(row, "Verdict_Fundamental",  "N/A"), row_get(row, "Score_Fund")),
            ("📉 Technicals",   row_get(row, "Verdict_Technical",    "N/A"), row_get(row, "Score_Tech")),
            ("📰 Sentiment",    row_get(row, "Sentiment_Label",      "N/A"), None),
        ],
    ):
        clr = get_color(str(verdict))
        score_txt = f"Score: {score}" if score is not None else ""
        col.markdown(
            f"<div style='text-align:center;border:1px solid #333;"
            f"border-radius:8px;padding:12px;'>"
            f"<b>{label}</b><br>"
            f"<span style='font-size:1.3em;color:{clr};'><b>{verdict}</b></span><br>"
            f"<small style='color:#888;'>{score_txt}</small></div>",
            unsafe_allow_html=True,
        )

    # Method breakdown tabs
    st.divider()
    st.subheader("📊 Method Breakdown")
    tab_val, tab_fund, tab_tech, tab_sent = st.tabs(
        ["💰 Valuation", "🏢 Fundamentals", "📉 Technicals", "📰 Sentiment"]
    )

    # ── VALUATION ─────────────────────────────────────────────────────────────
    with tab_val:
        bench = SECTOR_BENCHMARKS.get(str(row_get(row, "Sector", "")), DEFAULT_BENCHMARKS)
        v1, v2 = st.columns([1, 2])
        with v1:
            vv = row_get(row, "Verdict_Valuation", "N/A")
            st.markdown(f"<h3 style='color:{get_color(vv)};'>{vv}</h3>", unsafe_allow_html=True)
            st.caption("Multiples · Graham Number · DCF")
            dcf_v    = safe_num(row_get(row, "DCF_Value"))
            graham_v = safe_num(row_get(row, "Graham_Number"))
            price_v  = safe_num(row_get(row, "Price")) or 0
            ic1, ic2 = st.columns(2)
            if dcf_v and dcf_v > 0:
                ic1.metric("DCF Intrinsic Value", f"${dcf_v:,.2f}",
                           delta=f"${price_v - dcf_v:+,.2f} vs price",
                           delta_color="inverse")
                if detail.get("dcf_note"):
                    st.caption(detail["dcf_note"])
            if graham_v and graham_v > 0:
                ic2.metric("Graham Fair Value", f"${graham_v:,.2f}",
                           delta=f"${price_v - graham_v:+,.2f} vs price",
                           delta_color="inverse")
            if detail.get("val"):
                st.markdown("**Signals**")
                for s in detail["val"].get("reasons", []): st.markdown(f"- {s}")
        with v2:
            metrics = [
                ("P/E Ratio",   row_get(row,"PE_Ratio"),    bench["PE"],        "Trailing earnings multiple"),
                ("Forward P/E", row_get(row,"Forward_PE"),   None,               "Forward-looking multiple"),
                ("PEG Ratio",   row_get(row,"PEG_Ratio"),    1.0,                "< 1 = undervalued growth"),
                ("P/S Ratio",   row_get(row,"PS_Ratio"),     bench["PS"],        "Price to Sales (TTM)"),
                ("P/B Ratio",   row_get(row,"PB_Ratio"),     bench["PB"],        "Price to Book"),
                ("EV/EBITDA",   row_get(row,"EV_EBITDA"),    bench["EV_EBITDA"], "Enterprise value multiple"),
            ]
            st.dataframe(pd.DataFrame([{
                "Metric": m,
                "Stock":  fmt(sv, ".2f") if sv is not None else "N/A",
                "Sector Benchmark": f"{bv:.1f}" if isinstance(bv, float) else "–",
                "Note": note,
            } for m, sv, bv, note in metrics]), use_container_width=True, hide_index=True)

    # ── FUNDAMENTALS ──────────────────────────────────────────────────────────
    with tab_fund:
        f1, f2 = st.columns([1, 2])
        with f1:
            fv = row_get(row, "Verdict_Fundamental", "N/A")
            st.markdown(f"<h3 style='color:{get_color(fv)};'>{fv}</h3>", unsafe_allow_html=True)
            st.caption("ROE · Margins · Debt · Earnings Surprises")
            beat_rate = safe_num(row_get(row, "Earnings_Beat_Rate"))
            if beat_rate is not None:
                st.metric("EPS Beat Rate", f"{beat_rate*100:.0f}%",
                          help="% of last 4 quarters where actual EPS exceeded estimates")
            if detail.get("fund"):
                st.markdown("**Signals**")
                for s in detail["fund"].get("reasons", []): st.markdown(f"- {s}")
        with f2:
            fm1, fm2, fm3 = st.columns(3)
            fm1.metric("ROE",           fmt(row_get(row,"ROE"),            ".1f", suffix="%", scale=100), "Target: > 15%")
            fm2.metric("Profit Margin", fmt(row_get(row,"Profit_Margins"),  ".1f", suffix="%", scale=100), "Target: > 20%")
            fm3.metric("Debt / Equity", fmt(row_get(row,"Debt_to_Equity"),  ".0f", suffix="%"),             "Target: < 100%",
                       delta_color="inverse")

    # ── TECHNICALS ────────────────────────────────────────────────────────────
    with tab_tech:
        t1, t2 = st.columns([1, 2])
        with t1:
            tv = row_get(row, "Verdict_Technical", "N/A")
            st.markdown(f"<h3 style='color:{get_color(tv)};'>{tv}</h3>", unsafe_allow_html=True)
            st.caption("RSI · MACD · SMA · Volume · Beta")
            tm1, tm2, tm3, tm4 = st.columns(4)
            tm1.metric("RSI (14)",    fmt(row_get(row,"RSI"), ".1f"),    "30=Oversold · 70=Overbought")
            tm2.metric("200-SMA",     str(row_get(row,"SMA_Status","N/A")))
            tm3.metric("MACD",        str(row_get(row,"MACD_Signal","N/A")))
            beta_v = safe_num(row_get(row, "Beta"))
            tm4.metric("Beta",        f"{beta_v:.2f}" if beta_v is not None else "N/A",
                       help="Market risk relative to S&P 500. >1 = more volatile.")
            vol_sig = row_get(row, "Volume_Signal", "Normal")
            if vol_sig and vol_sig != "Normal":
                st.info(f"📊 Volume Signal: **{vol_sig}**")
            if detail.get("tech"):
                st.markdown("**Signals**")
                for s in detail["tech"].get("reasons", []): st.markdown(f"- {s}")
        with t2:
            hist_df = detail.get("hist")
            if hist_df is not None and not hist_df.empty:
                close = hist_df["Close"]
                fig   = go.Figure()
                fig.add_trace(go.Scatter(x=hist_df.index, y=close, name="Price",
                                         line=dict(color="#4FC3F7", width=1.5)))
                fig.add_trace(go.Scatter(x=hist_df.index, y=close.rolling(50).mean(),
                                         name="50-SMA", line=dict(color="#FFD54F",width=1,dash="dot")))
                fig.add_trace(go.Scatter(x=hist_df.index, y=close.rolling(200).mean(),
                                         name="200-SMA", line=dict(color="#EF5350",width=1,dash="dash")))
                fig.add_trace(go.Bar(x=hist_df.index, y=hist_df["Volume"], name="Volume",
                                     marker_color="rgba(100,100,200,0.25)", yaxis="y2"))
                fig.update_layout(
                    title=f"{ticker_input} — 1-Year Price + Volume",
                    xaxis_title="Date", yaxis_title="Price (USD)",
                    yaxis2=dict(overlaying="y", side="right", showgrid=False,
                                title="Volume", tickformat=".2s"),
                    legend=dict(orientation="h"), height=380,
                    margin=dict(l=0,r=0,t=40,b=0),
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#ccc"),
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Run a fresh analysis to display the price chart.")

    # ── SENTIMENT ─────────────────────────────────────────────────────────────
    with tab_sent:
        sent_score = safe_num(row_get(row, "Sentiment_Score"))
        sent_label = str(row_get(row, "Sentiment_Label", "N/A"))
        s1, s2 = st.columns([1, 2])
        with s1:
            st.markdown(f"<h3 style='color:{get_color(sent_label)};'>{sent_label}</h3>",
                        unsafe_allow_html=True)
            st.caption("News headlines scored with recency decay")
            if sent_score is not None:
                st.metric("Weighted Score", f"{sent_score:+.3f}",
                          help="−1 (max bearish) → +1 (max bullish). "
                               "Recent articles weighted higher via exponential decay.")
            if detail.get("sent"):
                st.markdown("**Summary**")
                for s in detail["sent"].get("reasons", []): st.markdown(f"- {s}")
        with s2:
            articles = detail.get("sent", {}).get("articles", [])
            if articles:
                st.markdown("**Scored Headlines** (newest first)")
                for art in sorted(articles, key=lambda x: x.get("age_days",99))[:12]:
                    age       = art.get("age_days","?")
                    decay_pct = art.get("decay", 1.0) * 100
                    st.markdown(
                        f"{art['signal']}  {art['title']}  "
                        f"<small style='color:#666;'>{age}d ago · decay {decay_pct:.0f}%</small>",
                        unsafe_allow_html=True,
                    )
            else:
                st.info("Run a fresh analysis to populate sentiment headlines.")

    # ── EXPORTS ───────────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📥 Export")
    ex1, ex2, ex3 = st.columns([1, 1, 2])

    with ex1:
        st.download_button(
            label=f"📄 {ticker_input}_analysis.csv",
            data=row.to_frame().T.to_csv(index=False).encode(),
            file_name=f"{ticker_input}_analysis.csv",
            mime="text/csv",
            use_container_width=True,
        )

    xlsx_key = f"{ticker_input}__{row_get(row,'Last_Updated','')}"
    if detail.get("info") and st.session_state.get("_xlsx_ticker") != xlsx_key:
        with st.spinner("Building Excel workbook…"):
            try:
                st.session_state["_xlsx_bytes"]  = generate_workbook(
                    ticker_input, row.to_dict(),
                    detail["info"], detail.get("hist", pd.DataFrame()), detail,
                )
                st.session_state["_xlsx_ticker"] = xlsx_key
            except Exception as exc:
                st.session_state["_xlsx_bytes"] = None
                st.warning(f"Could not build workbook: {exc}")

    xlsx_bytes = st.session_state.get("_xlsx_bytes")
    with ex2:
        if xlsx_bytes:
            st.download_button(
                label=f"📊 {ticker_input}_analysis.xlsx",
                data=xlsx_bytes,
                file_name=f"{ticker_input}_StockAnalysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
            )
        else:
            st.button("📊 Excel (run analysis first)", disabled=True, use_container_width=True)
    with ex3:
        st.caption(
            "**CSV** — raw analysis record from the database.  \n"
            "**Excel** — full research template pre-filled with 5 years of financials "
            "plus a styled Summary sheet. Requires a fresh analysis run."
        )


# =============================================================================
# TAB 2 — PORTFOLIO COMPARISON
# =============================================================================
with page_portfolio:
    st.subheader("📋 Watchlist — Side-by-Side Comparison")
    comp = db.get_watchlist_comparison()

    if comp.empty:
        st.info(
            "Your watchlist is empty. Analyse a ticker then click "
            "**☆ Add to Watchlist** to begin building your comparison."
        )
    else:
        def _verdict_bg(val):
            c = get_color(str(val)).lstrip("#")
            r_c, g_c, b_c = int(c[0:2],16), int(c[2:4],16), int(c[4:6],16)
            return (f"background-color: rgba({r_c},{g_c},{b_c},0.25); "
                    f"color: {get_color(str(val))}; font-weight: bold")

        display = comp[[
            "ticker","Price","Sector","Verdict_Overall",
            "Score_Val","Score_Fund","Score_Tech","Sentiment_Score",
            "PE_Ratio","PB_Ratio","EV_EBITDA",
            "DCF_Value","Graham_Number",
            "RSI","Beta","Earnings_Beat_Rate","Last_Updated",
        ]].copy()

        display.columns = [
            "Ticker","Price","Sector","Overall",
            "Val","Fund","Tech","Sent",
            "P/E","P/B","EV/EBITDA",
            "DCF","Graham",
            "RSI","Beta","EPS Beat %","Updated",
        ]
        for c in ["Price","P/E","P/B","EV/EBITDA","DCF","Graham","RSI","Beta"]:
            if c in display.columns:
                display[c] = pd.to_numeric(display[c], errors="coerce").round(2)
        if "Sent" in display.columns:
            display["Sent"] = pd.to_numeric(display["Sent"], errors="coerce").round(3)
        if "EPS Beat %" in display.columns:
            display["EPS Beat %"] = (
                pd.to_numeric(display["EPS Beat %"], errors="coerce") * 100
            ).round(0)

        st.dataframe(
            _styler_map(display.style, _verdict_bg, subset=["Overall"]),
            use_container_width=True, hide_index=True,
        )

        # Engine scores bar chart
        score_df = comp[["ticker","Score_Val","Score_Fund","Score_Tech"]].dropna()
        if not score_df.empty:
            score_df = score_df.copy()
            score_df.columns = ["Ticker","Valuation","Fundamentals","Technicals"]
            fig_p = px.bar(
                score_df.melt(id_vars="Ticker", var_name="Engine", value_name="Score"),
                x="Ticker", y="Score", color="Engine", barmode="group",
                title="Engine Scores — Watchlist Comparison",
                color_discrete_map={
                    "Valuation":    "#60A5FA",
                    "Fundamentals": "#34D399",
                    "Technicals":   "#FBBF24",
                },
            )
            fig_p.update_layout(
                height=320, margin=dict(l=0,r=0,t=40,b=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#ccc"),
            )
            st.plotly_chart(fig_p, use_container_width=True)

        st.download_button(
            "📄 Download comparison CSV",
            data=display.to_csv(index=False).encode(),
            file_name="watchlist_comparison.csv",
            mime="text/csv",
        )


# =============================================================================
# TAB 3 — SCORE HISTORY
# =============================================================================
with page_history:
    st.subheader(f"📈 Score History — {ticker_input}")
    hist_df = db.get_analysis_history(ticker_input)

    if hist_df.empty or len(hist_df) < 2:
        st.info(
            "Not enough history yet. Each **Run Full Analysis** saves a new snapshot. "
            "Run again on multiple occasions to see trends appear here."
        )
    else:
        hist_df["run_at"] = pd.to_datetime(hist_df["run_at"], errors="coerce")
        hist_df = hist_df.dropna(subset=["run_at"]).sort_values("run_at")

        # Score trend
        fig_s = go.Figure()
        for col, name, color in [
            ("Score_Val",  "Valuation",    "#60A5FA"),
            ("Score_Fund", "Fundamentals", "#34D399"),
            ("Score_Tech", "Technicals",   "#FBBF24"),
        ]:
            if col in hist_df.columns:
                fig_s.add_trace(go.Scatter(
                    x=hist_df["run_at"], y=hist_df[col],
                    name=name, mode="lines+markers",
                    line=dict(color=color, width=2), marker=dict(size=6),
                ))
        if "Sentiment_Score" in hist_df.columns:
            fig_s.add_trace(go.Scatter(
                x=hist_df["run_at"], y=hist_df["Sentiment_Score"],
                name="Sentiment", mode="lines+markers",
                line=dict(color="#F472B6", width=1.5, dash="dot"), marker=dict(size=5),
            ))
        fig_s.update_layout(
            title=f"{ticker_input} — Engine Scores Over Time",
            xaxis_title="Run Date", yaxis_title="Score",
            legend=dict(orientation="h"), height=360,
            margin=dict(l=0,r=0,t=40,b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#ccc"),
        )
        st.plotly_chart(fig_s, use_container_width=True)

        # Price vs intrinsic value
        if {"DCF_Value","Graham_Number","Price"}.issubset(hist_df.columns):
            fig_v = go.Figure()
            fig_v.add_trace(go.Scatter(x=hist_df["run_at"], y=hist_df["Price"],
                name="Market Price", line=dict(color="#4FC3F7", width=2)))
            fig_v.add_trace(go.Scatter(x=hist_df["run_at"], y=hist_df["DCF_Value"],
                name="DCF Intrinsic Value", line=dict(color="#34D399", width=2, dash="dash")))
            fig_v.add_trace(go.Scatter(x=hist_df["run_at"], y=hist_df["Graham_Number"],
                name="Graham Number", line=dict(color="#FBBF24", width=1.5, dash="dot")))
            fig_v.update_layout(
                title=f"{ticker_input} — Price vs Intrinsic Value Over Time",
                xaxis_title="Run Date", yaxis_title="USD",
                legend=dict(orientation="h"), height=320,
                margin=dict(l=0,r=0,t=40,b=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#ccc"),
            )
            st.plotly_chart(fig_v, use_container_width=True)

        # Verdict timeline table
        st.markdown("**Verdict Timeline**")
        if "Verdict_Overall" in hist_df.columns:
            vt_cols = [c for c in ["run_at","Verdict_Overall","Price",
                                    "Score_Val","Score_Fund","Score_Tech"]
                       if c in hist_df.columns]
            styled_hist = hist_df[vt_cols].rename(columns={
                "run_at":"Date","Verdict_Overall":"Verdict",
                "Score_Val":"Val","Score_Fund":"Fund","Score_Tech":"Tech",
            })
            st.dataframe(
                _styler_map(
                    styled_hist.style,
                    lambda v: f"color:{get_color(str(v))};font-weight:bold",
                    subset=["Verdict"],
                ),
                use_container_width=True, hide_index=True,
            )

        st.download_button(
            "📄 Download history CSV",
            data=hist_df.to_csv(index=False).encode(),
            file_name=f"{ticker_input}_score_history.csv",
            mime="text/csv",
        )


# =============================================================================
# TAB 4 — DATABASE & METHODOLOGY  (with SOX Certification)
# =============================================================================
with page_methodology:
    st.subheader("📚 Database & Methodology")

    # ── A. Database Overview ─────────────────────────────────────────────────
    with st.expander("🗄️ Database Schema & Storage", expanded=False):
        st.markdown(f"""
**Database file:** `{DB_FILENAME}` (SQLite 3, WAL journal mode)

| Table | Purpose |
|---|---|
| `analysis` | Latest snapshot per ticker — primary key `Ticker` |
| `analysis_history` | Append-only run log — every analysis appended, never replaced |
| `watchlist` | User-selected tickers for comparison |

**Key `analysis` columns:**

| Column | Type | Description |
|---|---|---|
| Ticker | TEXT | Exchange ticker symbol (e.g. AAPL) |
| Price | REAL | Closing price at time of analysis |
| Verdict_Overall | TEXT | Composite verdict (STRONG BUY → STRONG SELL) |
| DCF_Value | REAL | 5-year staged DCF intrinsic value per share |
| Graham_Number | REAL | √(22.5 × EPS × BVPS) |
| Score_Val / Fund / Tech | INTEGER | Raw engine sub-scores |
| Sentiment_Score | REAL | Recency-weighted headline score (−1 to +1) |
| Last_Updated | TEXT | YYYY-MM-DD HH:MM of last run |

The schema is **auto-migrated**: any new columns added to the codebase are  
automatically detected and added via `ALTER TABLE ADD COLUMN` on startup.
        """)

    # ── B. Scoring Methodology ───────────────────────────────────────────────
    with st.expander("⚙️ Scoring Methodology", expanded=False):
        st.markdown("""
### Composite Verdict Weights

| Engine | Weight | Verdict Range |
|---|---|---|
| Valuation | 35% | UNDERVALUED · FAIR VALUE · OVERVALUED |
| Fundamentals | 25% | STRONG · STABLE · WEAK |
| Technicals | 25% | BUY · HOLD · SELL |
| Sentiment | 15% | BULLISH → BEARISH |

Each engine maps to a signal of −1, 0, or +1. The weighted sum determines:

| Composite Score | Final Verdict |
|---|---|
| ≥ 0.55 | STRONG BUY |
| 0.20 – 0.54 | BUY |
| −0.19 – 0.19 | HOLD |
| −0.54 – −0.20 | SELL |
| ≤ −0.55 | STRONG SELL |

---

### Valuation Engine (35%)

Scores +1 per signal, −1 for penalties. Verdict threshold: score ≥ 3 → UNDERVALUED.

- **P/E vs sector benchmark** — sector-specific thresholds (e.g. Tech: 30×)  
- **EV/EBITDA vs sector benchmark**  
- **P/B vs sector benchmark**  
- **PEG ratio** — < 1.0 considered undervalued relative to growth  
- **Graham Number** = √(22.5 × trailing EPS × book value per share)  
- **DCF Intrinsic Value** — 5-year staged free cash flow model (see below); +2 for strong margin of safety ≥ 25%

#### DCF Model Parameters
- **Stage 1 growth (years 1–3):** analyst consensus earnings growth, capped ±25%  
- **Stage 2 growth (years 4–5):** 50% of Stage 1  
- **Terminal growth:** 2.5%  
- **WACC:** 4% risk-free + β × 5.5% equity risk premium, bounded [7%, 18%]  
- **Bridge:** EV → Equity = EV − total debt + cash → ÷ diluted shares

---

### Fundamental Engine (25%)

Scores +1/−1 per signal. Verdict threshold: score ≥ 3 → STRONG.

- **Return on Equity** > 15% ✅  
- **Net Profit Margin** > 20% ✅  
- **Debt/Equity** < 100% ✅, > 200% penalised  
- **Earnings surprise rate** — last 4 quarters: ≥ 75% beats = +2, ≥ 50% = +1, ≤ 25% = −1

---

### Technical Engine (25%)

Scores +1/−1 per signal. Verdict threshold: score ≥ 3 → BUY.

- **Price vs 200-SMA** (trend filter)  
- **Golden/Death Cross** (50 vs 200-SMA)  
- **RSI (14):** < 30 oversold (+2), > 70 overbought (−2)  
- **MACD (12/26/9):** line above signal = bullish  
- **Volume confirmation:** recent 5-day avg vs 20-day avg — breakout on >1.25× above SMA50 is +1; below SMA50 is −1  
- **Beta** displayed as context (no score impact)

---

### Sentiment Engine (15%)

- Headlines sourced from Yahoo Finance RSS (20–40 articles); falls back to yfinance feed  
- Each headline scored: positive-keyword count − negative-keyword count  
- Weighted by exponential time decay: weight = e^(−0.08 × age_days), half-life ≈ 8.7 days  
- Net weighted score clipped to [−1, +1]; label thresholds at ±0.15

---

### Data Source
All market data retrieved via **yfinance** (Yahoo Finance). Prices, financials,  
and fundamentals are as reported and subject to Yahoo Finance data quality.  
Each fetch includes 3-attempt retry logic with 2-second back-off.
        """)

    # ── C. SOX Certification ─────────────────────────────────────────────────
    st.divider()
    st.markdown("### 📋 Sarbanes-Oxley Style Certification")
    st.markdown(
        "The certification below is modelled on the disclosure requirements of the "
        "**Sarbanes-Oxley Act of 2002 (SOX), Sections 302 and 906**, adapted for an "
        "analytical tool rather than a public company filing."
    )

    with st.container(border=True):
        st.markdown("""
## CERTIFICATION OF ANALYTICAL DISCLOSURE

**Stock Engine Pro — Analysis Certification**  
Pursuant to the principles of the Sarbanes-Oxley Act of 2002, Sections 302 & 906

---

The undersigned Certifying Analyst, based on their knowledge, hereby certifies that:

**1. Accuracy of Analysis**  
The analysis outputs generated by this tool have been reviewed and, to the best of the
analyst's knowledge, accurately reflect the results of the disclosed computational
methodologies applied to the stated data inputs.

**2. Completeness of Disclosure**  
The methodology documentation contained in this tab fully describes all scoring engines,
weighting schemes, thresholds, and decision logic used to derive final verdicts. No material
methodology has been withheld or obscured.

**3. Data Sources Disclosed**  
All financial data is sourced from Yahoo Finance via the open-source `yfinance` library.
The analyst acknowledges that this data may contain errors, delays, or omissions, and that
results are only as reliable as the underlying data source.

**4. No Material Misrepresentations**  
The analysis does not knowingly contain any untrue statement of a material fact, nor does it
omit any material fact necessary to make the statements contained herein not misleading in
light of the circumstances under which such statements were made.

**5. Internal Controls Disclosure**  
- Retry logic (3 attempts, 2-second back-off) is applied to all external data calls  
- Data staleness is flagged when cache age exceeds **24 hours**  
- Ticker symbols are validated against Yahoo Finance before analysis commences  
- All computations use 64-bit floating point arithmetic with explicit `None`/`NaN` guards

**6. Limitations of Analysis**  
The analyst certifies that this tool is for **informational purposes only** and does **not**
constitute investment advice, a solicitation, or a recommendation to buy, hold, or sell any
security. Past model performance is not indicative of future results. The DCF and Graham Number
models incorporate assumptions that are inherently uncertain.

**7. No Conflict of Interest**  
This tool does not accept payment, commission, or any other consideration from any issuer,
broker, or financial institution in exchange for analysis outcomes.

---
        """)

        c1, c2, c3 = st.columns(3)
        with c1:
            analyst_name = st.text_input("Certifying Analyst Name", placeholder="Your name")
        with c2:
            analyst_title = st.text_input("Title / Role", placeholder="e.g. Portfolio Analyst")
        with c3:
            cert_date = st.date_input("Certification Date", value=datetime.date.today())

        if st.button("✅ Generate Signed Certification", type="primary"):
            if analyst_name.strip():
                st.success(
                    f"**Certified by:** {analyst_name}  |  "
                    f"**Title:** {analyst_title or 'N/A'}  |  "
                    f"**Date:** {cert_date.strftime('%B %d, %Y')}  \n\n"
                    "This certification has been recorded. The analyst acknowledges that knowingly "
                    "submitting false or misleading financial analysis may expose the preparer to "
                    "reputational, regulatory, and legal liability under applicable securities laws."
                )
                # Provide a downloadable text cert
                cert_text = f"""STOCK ENGINE PRO — ANALYTICAL CERTIFICATION
============================================================

Certifying Analyst : {analyst_name}
Title / Role       : {analyst_title or "N/A"}
Certification Date : {cert_date.strftime("%B %d, %Y")}
Ticker Analysed    : {ticker_input}
Analysis Timestamp : {row_get(row, "Last_Updated", "N/A")}

VERDICTS AT TIME OF CERTIFICATION
-----------------------------------
Overall Verdict  : {row_get(row, "Verdict_Overall",     "N/A")}
Valuation        : {row_get(row, "Verdict_Valuation",   "N/A")}  (score: {row_get(row, "Score_Val",  "N/A")})
Fundamentals     : {row_get(row, "Verdict_Fundamental", "N/A")}  (score: {row_get(row, "Score_Fund", "N/A")})
Technicals       : {row_get(row, "Verdict_Technical",   "N/A")}  (score: {row_get(row, "Score_Tech", "N/A")})
Sentiment        : {row_get(row, "Sentiment_Label",     "N/A")}  (score: {fmt(row_get(row, "Sentiment_Score"), "+.3f")})
Current Price    : ${row_get(row, "Price", 0):,.2f}
DCF Value        : {fmt(row_get(row, "DCF_Value"),    ".2f", prefix="$")}
Graham Number    : {fmt(row_get(row, "Graham_Number"), ".2f", prefix="$")}

CERTIFICATION STATEMENTS
-----------------------------------
The undersigned certifies that the analysis outputs above accurately reflect
the results of the disclosed methodologies documented in the Database &
Methodology tab of Stock Engine Pro. This analysis is for informational
purposes only and does not constitute investment advice.

Pursuant to the principles of the Sarbanes-Oxley Act of 2002, Sections 302 & 906,
the certifying analyst acknowledges that knowingly submitting false or misleading
financial analysis may expose the preparer to legal liability.

Signed: ______________________________
        {analyst_name}
        {analyst_title or "N/A"}
        {cert_date.strftime("%B %d, %Y")}

============================================================
Generated by Stock Engine Pro v4.1
Data source: Yahoo Finance (yfinance)
NOT INVESTMENT ADVICE
"""
                st.download_button(
                    "📄 Download Signed Certification (.txt)",
                    data=cert_text.encode(),
                    file_name=f"{ticker_input}_SOX_Certification_{cert_date.isoformat()}.txt",
                    mime="text/plain",
                )
            else:
                st.warning("Please enter the certifying analyst's name before signing.")

    st.caption(
        "⚠️ This certification is an analytical disclosure tool and does not constitute "
        "a legal filing under the Sarbanes-Oxley Act. It is designed to promote transparency "
        "and accountability in investment analysis."
    )
